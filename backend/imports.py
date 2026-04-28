"""Doctor list import — parses xlsx/csv and validates rows."""
import csv as _csv
import io
import re
from typing import Optional

ALLOWED_SEGMENTS = {"New", "Lapsed", "Occasional", "Active", "Engaged", "Expert"}
ALLOWED_DOCTOR_TYPES = {"GP", "Ortho", "Other"}
TARGET_FIELDS = ["first_name", "last_name", "doctor_name", "clinic_name", "city", "region", "doctor_type", "segment", "general_notes"]
REQUIRED_FIELDS = ["doctor_name"]   # but first+last together also satisfy this

# Default header → field aliases (case-insensitive, accent-insensitive after lowercasing)
DEFAULT_ALIASES = {
    "first_name": ["first_name", "first name", "firstname", "given name", "given", "fn", "name", "first"],
    "last_name": ["last_name", "last name", "lastname", "surname", "family name", "ln", "last"],
    "doctor_name": ["doctor_name", "doctor name", "doctor", "full name", "fullname", "physician", "dentist", "dr", "name"],
    "clinic_name": ["clinic_name", "clinic name", "clinic", "practice", "office", "facility"],
    "city": ["city", "town"],
    "region": ["region", "state", "province", "area", "country"],
    "doctor_type": ["doctor_type", "doctor type", "type", "specialty", "speciality"],
    "segment": ["segment", "tier", "level"],
    "general_notes": ["general_notes", "general notes", "notes", "comment", "comments", "remarks"],
}


def _norm(s: str) -> str:
    return re.sub(r"[\s_-]+", " ", (s or "").strip().lower())


def auto_map_headers(headers: list) -> dict:
    """Best-effort suggested mapping {target_field: header_label_or_None}."""
    mapping = {f: None for f in TARGET_FIELDS}
    norm_headers = {_norm(h): h for h in headers if h}
    for field, aliases in DEFAULT_ALIASES.items():
        for a in aliases:
            if _norm(a) in norm_headers:
                mapping[field] = norm_headers[_norm(a)]
                break
    return mapping


def parse_csv(blob: bytes) -> dict:
    """Returns {headers: [...], rows: [{header: cell, ...}]}."""
    text = blob.decode("utf-8-sig", errors="replace")
    reader = _csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return {"headers": [], "rows": []}
    headers = [str(h or "").strip() for h in rows[0]]
    out_rows = []
    for raw in rows[1:]:
        if not any((c or "").strip() for c in raw):
            continue
        d = {}
        for i, h in enumerate(headers):
            d[h] = (raw[i].strip() if i < len(raw) and raw[i] is not None else "")
        out_rows.append(d)
    return {"headers": headers, "rows": out_rows}


def parse_xlsx(blob: bytes) -> dict:
    """Read first worksheet of an xlsx file."""
    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(blob), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        first = next(rows_iter)
    except StopIteration:
        return {"headers": [], "rows": []}
    headers = [str(h).strip() if h is not None else "" for h in first]
    out_rows = []
    for raw in rows_iter:
        if not any((str(c).strip() if c is not None else "") for c in raw):
            continue
        d = {}
        for i, h in enumerate(headers):
            v = raw[i] if i < len(raw) else None
            d[h] = "" if v is None else str(v).strip()
        out_rows.append(d)
    return {"headers": headers, "rows": out_rows}


def parse_upload(filename: str, blob: bytes) -> dict:
    fname = (filename or "").lower()
    if fname.endswith(".xlsx"):
        return parse_xlsx(blob)
    if fname.endswith(".csv"):
        return parse_csv(blob)
    # Try csv fallback for unknown
    try:
        return parse_csv(blob)
    except Exception:
        return {"headers": [], "rows": []}


def validate_and_project(rows: list, mapping: dict) -> list:
    """Apply mapping to each row → project into target fields + collect per-row errors.

    If first_name and/or last_name are mapped, they are concatenated into doctor_name
    (preserving any explicitly mapped doctor_name when present and non-empty).

    Returns list of {row_index, raw, projected, errors[]}.
    """
    out = []
    for idx, raw in enumerate(rows):
        projected = {f: None for f in TARGET_FIELDS}
        for field, src in mapping.items():
            if not src:
                continue
            v = (raw.get(src) or "").strip() if isinstance(raw, dict) else ""
            projected[field] = v or None

        # Compose doctor_name from first/last when applicable
        first = (projected.get("first_name") or "").strip()
        last = (projected.get("last_name") or "").strip()
        existing_full = (projected.get("doctor_name") or "").strip()
        if not existing_full and (first or last):
            projected["doctor_name"] = f"{first} {last}".strip()

        errors = []
        if not projected.get("doctor_name"):
            errors.append("doctor_name is required (or first_name + last_name)")
        if projected.get("doctor_type"):
            t = projected["doctor_type"].strip().capitalize()
            if t.lower() in ("ortho", "orthodontist", "orthodontics"):
                t = "Ortho"
            elif t.lower() in ("gp", "general", "general practitioner", "dentist"):
                t = "GP"
            elif t not in ALLOWED_DOCTOR_TYPES:
                t = "Other"
            projected["doctor_type"] = t
        if projected.get("segment"):
            seg = projected["segment"].strip().capitalize()
            if seg not in ALLOWED_SEGMENTS:
                errors.append(f"segment must be one of {sorted(ALLOWED_SEGMENTS)}")
            else:
                projected["segment"] = seg
        out.append({
            "row_index": idx,
            "raw": raw,
            "projected": projected,
            "errors": errors,
        })
    return out


def template_rows() -> list:
    return [
        ["first_name", "last_name", "clinic_name", "city", "region", "doctor_type", "segment", "general_notes"],
        ["Ivan", "Ivanov", "Smile Clinic", "Sofia", "Sofia", "Ortho", "Active",
         "Interested in Invisalign but low clinical confidence"],
    ]


def build_template_csv() -> bytes:
    buf = io.StringIO()
    w = _csv.writer(buf)
    for r in template_rows():
        w.writerow(r)
    return buf.getvalue().encode("utf-8")


def build_template_xlsx() -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Doctors"
    rows = template_rows()
    for r in rows:
        ws.append(r)
    # Bold header
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    # Reasonable widths (8 columns now)
    widths = {"A": 14, "B": 16, "C": 24, "D": 16, "E": 16, "F": 14, "G": 14, "H": 50}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
